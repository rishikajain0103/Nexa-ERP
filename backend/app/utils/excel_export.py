import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def create_excel_file(filename, sheet_name, headers, rows):
    folder_path = "generated_reports"
    os.makedirs(folder_path, exist_ok=True)

    file_path = os.path.join(folder_path, filename)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_num, row_data in enumerate(rows, start=2):
        for col_num, value in enumerate(row_data, start=1):
            sheet.cell(row=row_num, column=col_num).value = value

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = max_length + 4

    workbook.save(file_path)

    return file_path